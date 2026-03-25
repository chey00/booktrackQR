# ------------------------------------------------------------------------------
# Projekt: BooktrackQR
# Modul: QR_Scanner (Fix: Relative Pfade & Super Clean)
# ------------------------------------------------------------------------------

import cv2
import csv
import os
import mysql.connector
import openpyxl
import re

# --- PFAD-FIX: Wir berechnen den Pfad relativ zur Datei ---
# os.path.dirname(__file__) gibt uns den Pfad zum "src" Ordner.
# Das ".." springt einen Ordner höher in das Hauptverzeichnis.
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX_PFAD = os.path.join(BASE_DIR, "schuelerlisten", "Testdaten Schüler.xlsx")
CSV_PFAD = os.path.join(BASE_DIR, "schuelerlisten", "Testdaten Schüler.csv")

DB_CONFIG = {
    'host': '192.168.10.195', 'port': 3306,
    'user': 'bookuser', 'password': '12345678', 'database': 'DB_BooktrackQR'
}


def super_clean(val):
    """Entfernt ALLES außer Buchstaben, Zahlen, Bindestriche und Unterstriche."""
    if val is None: return ""
    s = str(val).strip().upper()
    return re.sub(r'[^A-Z0-9\-_]', '', s)


def check_xlsx(scanned_id):
    if not os.path.exists(XLSX_PFAD):
        print(f"DEBUG: Excel Datei nicht gefunden unter {XLSX_PFAD}")
        return None
    try:
        wb = openpyxl.load_workbook(XLSX_PFAD, data_only=True)
        sheet = wb.active
        s_id = super_clean(scanned_id)

        # Finde Spalten
        headers = [super_clean(cell.value).lower() for cell in sheet[1]]
        id_col = -1
        for i, h in enumerate(headers):
            if "id" in h or "studierende" in h: id_col = i + 1

        if id_col != -1:
            for row in range(2, sheet.max_row + 1):
                raw_val = sheet.cell(row=row, column=id_col).value
                if super_clean(raw_val) == s_id:
                    # Namen/Infos aus der Zeile holen
                    vn = sheet.cell(row=row, column=2).value or ""
                    nn = sheet.cell(row=row, column=3).value or ""
                    return f"Excel: {vn} {nn}"
    except Exception as e:
        print(f"Excel-Error: {e}")
    return None


def check_csv(scanned_id):
    if not os.path.exists(CSV_PFAD):
        print(f"DEBUG: CSV Datei nicht gefunden unter {CSV_PFAD}")
        return None
    s_id = super_clean(scanned_id)
    try:
        with open(CSV_PFAD, mode='r', encoding='utf-8-sig') as f:
            lines = f.readlines()
            header = lines[0]
            sep = ";" if ";" in header else ","

            reader = csv.DictReader(lines, delimiter=sep)
            for row in reader:
                for key, val in row.items():
                    if super_clean(val) == s_id:
                        vn = row.get('vorname') or row.get('Vorname') or ""
                        nn = row.get('nachname') or row.get('Nachname') or ""
                        return f"CSV: {vn} {nn}"
    except Exception as e:
        print(f"CSV-Error: {e}")
    return None


def run_standalone():
    # Kurzer Check beim Start im Terminal
    print(f"--- Suche Dateien in: {BASE_DIR} ---")
    print(f"Excel vorhanden: {os.path.exists(XLSX_PFAD)}")
    print(f"CSV vorhanden: {os.path.exists(CSV_PFAD)}")

    cap = cv2.VideoCapture(0)
    detector = cv2.QRCodeDetector()

    while True:
        ret, frame = cap.read()
        if not ret: break
        data, _, _ = detector.detectAndDecode(frame)

        if data:
            # Wir zeigen im Terminal exakt an, was die Kamera sieht
            print(f"Gescannter Inhalt: '{data}'")

            xlsx_res = check_xlsx(data)
            csv_res = check_csv(data)

            if xlsx_res or csv_res:
                msg = xlsx_res if xlsx_res else csv_res
                cv2.putText(frame, f"GEFUNDEN: {msg}", (20, 100), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 3)
            else:
                cv2.putText(frame, f"NICHT GEFUNDEN: {data}", (20, 100), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 0, 255), 2)

        cv2.imshow("Scanner Test", frame)
        if cv2.waitKey(1) & 0xFF == ord('q'): break

    cap.release()
    cv2.destroyAllWindows()


if __name__ == "__main__":
    run_standalone()